"""
CLI Interface - Consolidate historical reports into a single XLSX.

This module reads all JSON reports from reports_history/, deduplicates
packages by (name, version), and generates a single consolidated.xlsx
with conditional formatting for approval status.
"""

from __future__ import annotations
import json
import sys
from pathlib import Path
from typing import Dict, List, Any, Set, Tuple
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


class ConsolidationError(Exception):
    """Raised when consolidation process encounters an error."""


def find_all_report_jsons(base_directory: str | Path = "reports_history") -> List[Path]:
    """
    Find all consolidated_report.json files in the history directory.
    
    Returns:
        List of paths to JSON report files
        
    Raises:
        ConsolidationError: If directory doesn't exist or no reports found
    """
    base_dir = Path(base_directory)
    
    if not base_dir.exists():
        raise ConsolidationError(
            f"Reports history directory not found: {base_dir}\n"
            "Run at least one analysis first to generate reports."
        )
    
    # Find all consolidated_report.json files in timestamped subdirectories
    json_files = list(base_dir.glob("*/consolidated_report.json"))
    
    if not json_files:
        raise ConsolidationError(
            f"No report files found in {base_dir}\n"
            "Run at least one analysis first to generate reports."
        )
    
    return sorted(json_files)


def load_and_merge_packages(json_paths: List[Path]) -> Dict[Tuple[str, str], Dict[str, Any]]:
    """
    Load all JSON reports and merge packages, deduplicating by (name, version).
    
    Args:
        json_paths: List of paths to consolidated_report.json files
        
    Returns:
        Dictionary mapping (name, version) -> package data
        
    Raises:
        ConsolidationError: If JSON parsing fails
    """
    merged_packages: Dict[Tuple[str, str], Dict[str, Any]] = {}
    
    for json_path in json_paths:
        try:
            with json_path.open("r", encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError) as e:
            raise ConsolidationError(f"Failed to read {json_path}: {e}")
        
        packages = data.get("packages", [])
        
        for pkg in packages:
            name = pkg.get("package", "").strip()
            version = pkg.get("version", "").strip()
            
            if not name or not version:
                continue
            
            # Use (name, version) as unique key
            key = (name.lower(), version)
            
            # Keep first occurrence (or update with latest if preferred)
            if key not in merged_packages:
                merged_packages[key] = pkg
    
    return merged_packages


def generate_consolidated_xlsx(
    packages: Dict[Tuple[str, str], Dict[str, Any]],
    output_path: str | Path = "consolidated.xlsx",
) -> None:
    """
    Generate consolidated XLSX with conditional formatting.
    
    Approved packages get pastel blue in "Aprobada" column.
    Rejected packages get pastel orange in "Aprobada" column.
    
    Args:
        packages: Dictionary of deduplicated packages
        output_path: Path to output XLSX file
        
    Raises:
        ConsolidationError: If XLSX generation fails
    """
    if not packages:
        raise ConsolidationError("No packages to consolidate")
    
    wb = Workbook()
    ws = wb.active
    
    if ws is None:
        raise ConsolidationError("Failed to create worksheet")
    
    ws.title = "Consolidated Packages"
    
    # Define headers (same as regular XLSX report)
    headers = [
        "Nombre", "Versión", "Licencia", "Aprobada", "Estado / Comentario",
        "Dependencias Directas", "Dependencias Transitivas", "Dependencias Rechazadas",
        "Fecha de Publicación", "URL", "Descripción"
    ]
    ws.append(headers)
    
    # Style header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Define fills for conditional formatting
    approved_fill = PatternFill(
        start_color="D9E1F2",  # Pastel blue
        end_color="D9E1F2",
        fill_type="solid",
    )
    rejected_fill = PatternFill(
        start_color="FFE699",  # Pastel orange
        end_color="FFE699",
        fill_type="solid",
    )
    
    # Sort packages by name for readability
    sorted_packages = sorted(packages.values(), key=lambda p: (p.get("package", "").lower(), p.get("version", "")))
    
    # Helper to safely extract values
    def safe_value(val: Any) -> str:
        """Convert value to safe string."""
        if val is None or val == "":
            return "N/A"
        val_str = str(val).strip()
        return val_str if val_str else "N/A"
    
    def format_deps(deps: Any) -> str:
        """Format dependency list."""
        if not deps:
            return "N/A"
        if isinstance(deps, list):
            dep_names = []
            for dep in deps:
                if isinstance(dep, dict):
                    dep_names.append(dep.get("name", "?"))
                elif isinstance(dep, str):
                    dep_names.append(dep)
            return ", ".join(dep_names) if dep_names else "N/A"
        return safe_value(deps)
    
    # Write package rows
    for pkg in sorted_packages:
        name = safe_value(pkg.get("package", ""))
        version = safe_value(pkg.get("version", ""))
        license_info = safe_value(pkg.get("licencia") or pkg.get("license", ""))
        
        # Handle both boolean and string formats for approval status
        approved_raw = pkg.get("aprobada", False)
        if isinstance(approved_raw, str):
            approved = approved_raw.strip().lower() in ("sí", "si", "yes", "true")
        else:
            approved = bool(approved_raw)
        
        rejection_reason = safe_value(pkg.get("motivo_rechazo", ""))
        
        # Dependencies
        dep_directas = format_deps(pkg.get("dependencias_directas", []))
        dep_transitivas = format_deps(pkg.get("dependencias_transitivas", []))
        dep_rechazadas = format_deps(pkg.get("dependencias_rechazadas", []))
        
        # Date formatting
        fecha_pub = pkg.get("upload_time") or pkg.get("upload_time_iso_8601") or ""
        if isinstance(fecha_pub, str) and "T" in fecha_pub:
            fecha_pub = fecha_pub.split("T")[0]
        fecha_pub = safe_value(fecha_pub)
        
        # URL
        url = pkg.get("project_url") or pkg.get("home_page") or (f"https://pypi.org/project/{name}" if name != "N/A" else "")
        url = safe_value(url)
        
        # Description
        descripcion = pkg.get("summary") or pkg.get("description") or ""
        if isinstance(descripcion, str):
            descripcion = " ".join(descripcion.split())
            if len(descripcion) > 200:
                descripcion = descripcion[:197] + "..."
        descripcion = safe_value(descripcion)
        
        # Convert boolean to Spanish string
        approved_text = "Sí" if approved else "No"
        
        row_data: List[str] = [
            name, version, license_info, approved_text, rejection_reason,
            dep_directas, dep_transitivas, dep_rechazadas,
            fecha_pub, url, descripcion
        ]
        ws.append(row_data)
        
        # Apply conditional formatting to "Aprobada" column (column 4)
        current_row = ws.max_row
        aprobada_cell = ws.cell(row=current_row, column=4)
        
        if approved:
            aprobada_cell.fill = approved_fill
        else:
            aprobada_cell.fill = rejected_fill
    
    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
    
    try:
        wb.save(output_path)
    except OSError as e:
        raise ConsolidationError(f"Failed to save XLSX file: {e}")


def main() -> None:
    """
    Main entry point for consolidation CLI.
    
    Reads all historical reports, deduplicates packages,
    and generates a single consolidated.xlsx file.
    """
    try:
        print("[CONSOLIDATE] Finding historical reports...")
        json_paths = find_all_report_jsons()
        print(f"[CONSOLIDATE] Found {len(json_paths)} report(s)")
        
        print("[CONSOLIDATE] Loading and merging packages...")
        packages = load_and_merge_packages(json_paths)
        print(f"[CONSOLIDATE] Deduplicated to {len(packages)} unique package(s)")
        
        output_file = "consolidated.xlsx"
        print(f"[CONSOLIDATE] Generating {output_file}...")
        generate_consolidated_xlsx(packages, output_file)
        
        print(f"[OK] Consolidated report generated: {output_file}")
        sys.exit(0)
        
    except ConsolidationError as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)
        
    except Exception as e:
        print(f"[ERROR] Unexpected error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
