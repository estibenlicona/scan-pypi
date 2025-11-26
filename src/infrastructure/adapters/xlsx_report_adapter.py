"""XLSX Report Adapter - Generates Excel .xlsx with styling."""

import json
from pathlib import Path
from typing import Any, List, Optional, Dict
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment
from typing import cast
from src.domain.ports import LoggerPort


class XLSXReportAdapter:
    """Generates an Excel .xlsx file with package metadata and styling."""

    def __init__(self, logger: LoggerPort) -> None:
        self.logger = logger
    
    def _safe_cell_value(self, value: Any) -> str:
        """Convert value to safe string for Excel cell."""
        if value is None or value == "":
            return "N/A"
        value_str = str(value).strip()
        if value_str.startswith(("=", "+", "-", "@")):
            value_str = "'" + value_str
        value_str = value_str.replace("—", "-").replace("–", "-")
        return value_str if value_str else "N/A"
    
    def _get_root_fill(self) -> PatternFill:
        """Get pastel blue fill for root libraries."""
        return PatternFill(start_color="CCE8FF", end_color="CCE8FF", fill_type="solid")
    
    def _get_rejected_fill(self) -> PatternFill:
        """Get pastel red fill for rejected packages."""
        return PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    def _apply_row_styling(self, ws: Worksheet, row_num: int, is_root: bool, is_rejected: bool) -> None:
        """Apply styling to main columns."""
        main_columns = [1, 2, 3, 4]
        if is_rejected:
            fill = self._get_rejected_fill()
        elif is_root:
            fill = self._get_root_fill()
        else:
            return
        for col_idx in main_columns:
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical='center')
    
    def _format_dependencies(self, deps_list: List[Any]) -> str:
        """Format dependency list."""
        if not deps_list:
            return "N/A"
        dep_names = []
        for dep in deps_list:
            if isinstance(dep, dict):
                dep_names.append(dep.get("name", "?"))
            elif isinstance(dep, str):
                dep_names.append(dep)
        return ", ".join(dep_names) if dep_names else "N/A"
    
    def _sort_packages_by_root_libraries(self, packages: List[Any], root_libraries: Optional[List[str]]) -> List[Any]:
        """Sort packages with root first."""
        if not root_libraries:
            return packages
        root_libs_lower = [lib.lower() for lib in root_libraries]
        root_libs_set = set(root_libs_lower)
        root_packages = []
        other_packages = []
        for pkg in packages:
            pkg_name = pkg.get("package", "").lower()
            if pkg_name in root_libs_set:
                root_packages.append(pkg)
            else:
                other_packages.append(pkg)
        def get_root_order(pkg: Dict[str, Any]) -> int:
            pkg_name = pkg.get("package", "").lower()
            try:
                return root_libs_lower.index(pkg_name)
            except ValueError:
                return len(root_libs_lower)
        root_packages.sort(key=get_root_order)
        sorted_packages = root_packages + other_packages
        if root_packages:
            self.logger.info(f"Ordered {len(root_packages)} root libraries first, {len(other_packages)} dependencies")
        return sorted_packages

    def generate_xlsx(self, report_path: str, output_path: str = "packages.xlsx", root_libraries: Optional[List[str]] = None) -> bool:
        """Generate XLSX report."""
        try:
            if not Path(report_path).exists():
                self.logger.error(f"Report not found: {report_path}")
                return False

            with open(report_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            packages = data.get("packages", [])
            if not packages:
                self.logger.warning("No packages found")
                return False

            # Deduplicate
            seen: Dict[str, bool] = {}
            deduplicated_packages = []
            for pkg in packages:
                key = f"{pkg.get('package', '')}@{pkg.get('version', '')}"
                if key not in seen:
                    seen[key] = True
                    deduplicated_packages.append(pkg)
            if len(deduplicated_packages) < len(packages):
                self.logger.info(f"Removed {len(packages) - len(deduplicated_packages)} duplicates")
            packages = deduplicated_packages
            
            # Sort
            packages = self._sort_packages_by_root_libraries(packages, root_libraries)

            # Create workbook
            wb = Workbook()
            ws_or_none = wb.active
            if ws_or_none is None:
                ws_or_none = wb.create_sheet()
            ws: Worksheet = cast(Worksheet, ws_or_none)
            ws.title = "Packages"
            header = [
                "Nombre", "Versión", "Licencia", "Aprobada", "Estado / Comentario",
                "Dependencias Directas", "Dependencias Transitivas", "Dependencias Rechazadas",
                "Fecha de Publicación", "URL", "Descripción"
            ]
            ws.append(header)

            # Normalize root libraries
            root_libs_lower = [lib.lower() for lib in (root_libraries or [])]
            root_libs_set = set(root_libs_lower)

            for row_idx, pkg in enumerate(packages, start=2):
                nombre = self._safe_cell_value(pkg.get("package") or "")
                version = self._safe_cell_value(pkg.get("version") or "")
                licencia = self._safe_cell_value(pkg.get("license") or "")
                aprobada = self._safe_cell_value(pkg.get("aprobada", ""))
                estado_comentario = self._safe_cell_value(pkg.get("motivo_rechazo") or "")
                
                dependencias_directas_list = pkg.get("dependencias_directas", [])
                dependencias_transitivas_list = pkg.get("dependencias_transitivas", [])
                dependencias_rechazadas_list = pkg.get("dependencias_rechazadas", [])
                
                dep_directas_str = self._safe_cell_value(self._format_dependencies(dependencias_directas_list))
                dep_transitivas_str = self._safe_cell_value(self._format_dependencies(dependencias_transitivas_list))
                dep_rechazadas_str = self._safe_cell_value(self._format_dependencies(dependencias_rechazadas_list))
                
                fecha_pub = pkg.get("upload_time") or pkg.get("upload_time_iso_8601") or ""
                if isinstance(fecha_pub, str) and "T" in fecha_pub:
                    fecha_pub = fecha_pub.split("T")[0]
                fecha_pub = self._safe_cell_value(fecha_pub)
                
                url = pkg.get("project_url") or pkg.get("home_page") or (f"https://pypi.org/project/{nombre}" if nombre and nombre != "N/A" else "")
                url = self._safe_cell_value(url)
                
                descripcion = pkg.get("summary") or pkg.get("description") or ""
                if isinstance(descripcion, str):
                    descripcion = " ".join(descripcion.split())
                    if len(descripcion) > 200:
                        descripcion = descripcion[:197] + "..."
                descripcion = self._safe_cell_value(descripcion)
                
                row: List[str] = [
                    nombre, version, licencia, aprobada, estado_comentario,
                    dep_directas_str, dep_transitivas_str, dep_rechazadas_str,
                    fecha_pub, url, descripcion
                ]
                ws.append(row)
                
                # Apply styling
                is_root = nombre.lower() in root_libs_set
                is_rejected = aprobada.strip().lower() == "no"
                self._apply_row_styling(ws, row_idx, is_root, is_rejected)

            # Auto-size
            for idx, col in enumerate(ws.columns, 1):
                max_length = 0
                col_letter = get_column_letter(idx)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

            wb.save(output_path)
            self.logger.info(f"✅ XLSX report generated: {output_path}")
            return True

        except Exception as e:
            self.logger.error(f"Failed to generate XLSX: {e}")
            return False
