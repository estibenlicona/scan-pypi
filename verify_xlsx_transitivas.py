import openpyxl

wb = openpyxl.load_workbook('packages.xlsx')
ws = wb.active

headers = [str(cell.value) for cell in ws[1]]
print("Headers:", headers)

try:
    directas_idx = headers.index('Dependencias Directas')
    transitivas_idx = headers.index('Dependencias Transitivas')
    pkg_idx = headers.index('Nombre')
    
    print(f"\nLooking for ipykernel and ipython:\n")
    
    for i in range(2, min(200, ws.max_row+1)):
        pkg_name = ws.cell(i, pkg_idx+1).value
        if pkg_name in ['ipykernel', 'ipython']:
            directas = str(ws.cell(i, directas_idx+1).value or "")
            transitivas = str(ws.cell(i, transitivas_idx+1).value or "")
            
            print(f"{pkg_name}:")
            print(f"  Directas: {directas[:100]}")
            print(f"  Transitivas: {transitivas[:100]}")
            print()
            
            # Check if colorama appears in the right place
            if pkg_name == 'ipykernel':
                colorama_in_directas = 'colorama' in directas.lower()
                colorama_in_transitivas = 'colorama' in transitivas.lower()
                print(f"  ✓ colorama in directas: {colorama_in_directas} (should be False)")
                print(f"  ✓ colorama in transitivas: {colorama_in_transitivas} (should be True)")
            
except ValueError as e:
    print(f"Column not found: {e}")
