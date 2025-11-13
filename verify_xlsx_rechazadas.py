import openpyxl

wb = openpyxl.load_workbook('packages.xlsx')
ws = wb.active

headers = [cell.value for cell in ws[1]]
print("Headers:", headers)

try:
    idx = headers.index('Dependencias Rechazadas')
    print(f"\nColumn 'Dependencias Rechazadas' found at index: {idx}")
    print("\nSample rows with rejected dependencies:")
    count = 0
    for i in range(2, min(20, ws.max_row+1)):
        value = ws.cell(i, idx+1).value
        pkg_name = ws.cell(i, 1).value
        if value and value != "N/A":
            print(f"  {pkg_name}: {value}")
            count += 1
            if count >= 5:
                break
except ValueError:
    print("Column 'Dependencias Rechazadas' not found!")
