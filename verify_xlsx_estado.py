import openpyxl

wb = openpyxl.load_workbook('packages.xlsx')
ws = wb.active

headers = [str(cell.value) for cell in ws[1]]

try:
    pkg_idx = headers.index('Nombre')
    estado_idx = headers.index('Estado / Comentario')
    rechazadas_idx = headers.index('Dependencias Rechazadas')
    
    print("Packages with rejection reason:\n")
    
    for i in range(2, min(50, ws.max_row+1)):
        pkg_name = ws.cell(i, pkg_idx+1).value
        estado = str(ws.cell(i, estado_idx+1).value or "")
        rechazadas = str(ws.cell(i, rechazadas_idx+1).value or "")
        
        if pkg_name and 'Dependencias rechazadas' in estado:
            print(f"{pkg_name}:")
            print(f"  Estado: {estado}")
            print(f"  Rechazadas: {rechazadas}")
            print()
            
except ValueError as e:
    print(f"Column not found: {e}")
