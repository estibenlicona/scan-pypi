#!/usr/bin/env python3
"""Test script to verify package sorting works correctly."""

import json
from datetime import datetime

# Create test consolidated report with packages in random order
test_report = {
    "timestamp": datetime.now().isoformat(),
    "vulnerabilities": [],
    "packages": [
        {"package": "scipy", "version": "1.10.0", "license": "BSD"},
        {"package": "numpy", "version": "1.24.0", "license": "BSD"},
        {"package": "matplotlib", "version": "3.7.0", "license": "PSF"},
        {"package": "pandas", "version": "2.0.0", "license": "BSD"},
        {"package": "requests", "version": "2.31.0", "license": "Apache-2.0"},
        {"package": "sklearn", "version": "1.3.0", "license": "BSD"},
    ],
    "filtered_packages": []
}

# Save test report
with open("test_sorting_report.json", "w") as f:
    json.dump(test_report, f, indent=2)

print("✅ Test report created with 6 packages in random order")

# Define root libraries in specific order
root_libraries = ["numpy", "pandas", "matplotlib", "requests"]

print(f"\n📋 Root libraries (in order):")
for i, lib in enumerate(root_libraries, 1):
    print(f"   {i}. {lib}")

# Generate XLSX with ordering
from src.infrastructure.adapters.xlsx_report_adapter import XLSXReportAdapter
from src.infrastructure.adapters.logger_adapter import LoggerAdapter
from src.infrastructure.config.settings import LoggingSettings

logger = LoggerAdapter(LoggingSettings())
xlsx_adapter = XLSXReportAdapter(logger)

# Generate XLSX from test report with sorting
if xlsx_adapter.generate_xlsx("test_sorting_report.json", "test_sorting_packages.xlsx", root_libraries=root_libraries):
    print("\n✅ XLSX generated successfully")
    
    # Now let's verify the order in the Excel file
    from openpyxl import load_workbook
    wb = load_workbook("test_sorting_packages.xlsx")
    ws = wb.active
    
    print("\n📊 Package order in Excel (Row = Package):")
    print("-" * 50)
    
    package_order = []
    for row_idx in range(2, ws.max_row + 1):  # Skip header
        package_name = ws[f"A{row_idx}"].value
        if package_name:
            package_order.append(package_name)
            print(f"   Row {row_idx}: {package_name}")
    
    print("\n✅ VERIFICATION:")
    print("-" * 50)
    
    # Check if root libraries appear first in correct order
    root_in_excel = package_order[:len(root_libraries)]
    print(f"\n   Expected order (root first): {root_libraries}")
    print(f"   Actual order in Excel:       {root_in_excel}")
    
    if root_in_excel == root_libraries:
        print("\n   ✅ ROOT LIBRARIES CORRECTLY ORDERED!")
    else:
        print("\n   ❌ Root libraries NOT in correct order")
    
    # Check remaining packages
    remaining = package_order[len(root_libraries):]
    if remaining:
        print(f"\n   Other packages: {remaining}")
    
    print("\n✅ Test completed successfully!")
else:
    print("❌ Failed to generate XLSX")
