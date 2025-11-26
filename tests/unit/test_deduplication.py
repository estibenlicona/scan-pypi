#!/usr/bin/env python3
"""Test script to verify package deduplication works correctly."""

import json
from datetime import datetime

# Create test consolidated report with duplicate packages
test_report = {
    "timestamp": datetime.now().isoformat(),
    "vulnerabilities": [],
    "packages": [
        {"package": "numpy", "version": "1.24.0", "license": "BSD"},
        {"package": "pandas", "version": "2.0.0", "license": "BSD"},
        {"package": "numpy", "version": "1.24.0", "license": "BSD"},  # Duplicate
        {"package": "scipy", "version": "1.10.0", "license": "BSD"},
        {"package": "pandas", "version": "2.0.0", "license": "BSD"},  # Duplicate
        {"package": "matplotlib", "version": "3.7.0", "license": "PSF"},
        {"package": "numpy", "version": "1.24.0", "license": "BSD"},  # Another duplicate
    ],
    "filtered_packages": []
}

# Save test report
with open("test_report.json", "w") as f:
    json.dump(test_report, f, indent=2)

print(f"✅ Test report created with {len(test_report['packages'])} packages (including duplicates)")
print(f"   Unique packages should be: 4 (numpy, pandas, scipy, matplotlib)")

# Now test the deduplication logic
from src.infrastructure.adapters.xlsx_report_adapter import XLSXReportAdapter
from src.infrastructure.adapters.logger_adapter import LoggerAdapter
from src.infrastructure.config.settings import LoggingSettings

logger = LoggerAdapter(LoggingSettings())
xlsx_adapter = XLSXReportAdapter(logger)

# Generate XLSX from test report
if xlsx_adapter.generate_xlsx("test_report.json", "test_packages.xlsx"):
    print("✅ XLSX generated successfully")
    
    # Now let's count the rows in the Excel file
    from openpyxl import load_workbook
    wb = load_workbook("test_packages.xlsx")
    ws = wb.active
    
    # Count data rows (excluding header)
    data_rows = ws.max_row - 1
    print(f"✅ XLSX contains {data_rows} data rows (header + data)")
    print(f"   Expected: 4 data rows (1 header + 4 unique packages)")
    
    if data_rows == 4:
        print("✅ DEDUPLICATION WORKING CORRECTLY!")
    else:
        print(f"❌ DEDUPLICATION FAILED: Expected 4 rows but got {data_rows}")
else:
    print("❌ Failed to generate XLSX")

print("\n✅ Test completed!")
